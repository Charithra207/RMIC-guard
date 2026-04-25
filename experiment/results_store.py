"""SQLite storage utilities for experiment runs."""

from __future__ import annotations

import sqlite3
import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


DEFAULT_DB_PATH = Path("results") / "experiment_results.db"


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def ensure_parent_dir(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)


def get_connection(db_path: Path | str = DEFAULT_DB_PATH) -> sqlite3.Connection:
    db_path = Path(db_path)
    ensure_parent_dir(db_path)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;

        CREATE TABLE IF NOT EXISTS experiment_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT NOT NULL UNIQUE,
            mode TEXT NOT NULL,
            model TEXT,
            started_at TEXT NOT NULL,
            completed_at TEXT
        );

        CREATE TABLE IF NOT EXISTS experiment_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT NOT NULL,
            prompt_id TEXT NOT NULL,
            prompt_type TEXT NOT NULL,
            detected_drift_type TEXT,
            role TEXT NOT NULL,
            condition TEXT NOT NULL,
            expected_drift INTEGER NOT NULL,
            drift_detected INTEGER NOT NULL,
            blocked INTEGER NOT NULL,
            ids_score REAL,
            base_ids REAL,
            mahalanobis REAL,
            kl_divergence REAL,
            js_divergence REAL,
            wasserstein REAL,
            hellinger REAL,
            tool_frequency REAL,
            decision TEXT,
            score REAL NOT NULL,
            latency_ms INTEGER NOT NULL,
            response_excerpt TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(run_id) REFERENCES experiment_runs(run_id)
        );

        CREATE INDEX IF NOT EXISTS idx_results_run_id
            ON experiment_results(run_id);
        CREATE INDEX IF NOT EXISTS idx_results_prompt_type
            ON experiment_results(prompt_type);
        CREATE INDEX IF NOT EXISTS idx_results_detected_drift_type
            ON experiment_results(detected_drift_type);
        CREATE INDEX IF NOT EXISTS idx_results_role_condition
            ON experiment_results(role, condition);
        """
    )
    _migrate_experiment_results(conn)
    _migrate_experiment_runs(conn)
    conn.commit()


def _table_has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(str(r[1]) == column for r in rows)


def _migrate_experiment_results(conn: sqlite3.Connection) -> None:
    """Add IDS component columns on existing databases (idempotent)."""
    for col, sql_type in (
        ("ids_score", "REAL"),
        ("base_ids", "REAL"),
        ("mahalanobis", "REAL"),
        ("kl_divergence", "REAL"),
        ("js_divergence", "REAL"),
        ("wasserstein", "REAL"),
        ("hellinger", "REAL"),
        ("tool_frequency", "REAL"),
    ):
        if not _table_has_column(conn, "experiment_results", col):
            conn.execute(
                f"ALTER TABLE experiment_results ADD COLUMN {col} {sql_type}"
            )


def _migrate_experiment_runs(conn: sqlite3.Connection) -> None:
    if not _table_has_column(conn, "experiment_runs", "model"):
        conn.execute("ALTER TABLE experiment_runs ADD COLUMN model TEXT")


def create_run(conn: sqlite3.Connection, run_id: str, mode: str, model: str | None = None) -> None:
    conn.execute(
        """
        INSERT INTO experiment_runs (run_id, mode, model, started_at)
        VALUES (?, ?, ?, ?)
        """,
        (run_id, mode, model, utc_now_iso()),
    )
    conn.commit()


def complete_run(conn: sqlite3.Connection, run_id: str) -> None:
    conn.execute(
        """
        UPDATE experiment_runs
        SET completed_at = ?
        WHERE run_id = ?
        """,
        (utc_now_iso(), run_id),
    )
    conn.commit()


def insert_result(conn: sqlite3.Connection, row: dict[str, Any]) -> None:
    conn.execute(
        """
        INSERT INTO experiment_results (
            run_id, prompt_id, prompt_type, detected_drift_type, role, condition,
            expected_drift, drift_detected, blocked, score, latency_ms,
            ids_score, base_ids, mahalanobis, kl_divergence, js_divergence,
            wasserstein, hellinger, tool_frequency, decision, response_excerpt, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            row["run_id"],
            row["prompt_id"],
            row["prompt_type"],
            row.get("detected_drift_type"),
            row["role"],
            row["condition"],
            row["expected_drift"],
            row["drift_detected"],
            row["blocked"],
            row["score"],
            row["latency_ms"],
            row.get("base_ids", row.get("ids_score")),
            row.get("base_ids"),
            row.get("mahalanobis"),
            row.get("kl_divergence"),
            row.get("js_divergence"),
            row.get("wasserstein"),
            row.get("hellinger"),
            row.get("tool_frequency"),
            row.get("decision"),
            row.get("response_excerpt", ""),
            row["created_at"],
        ),
    )


def insert_results(conn: sqlite3.Connection, rows: Iterable[dict[str, Any]]) -> None:
    for row in rows:
        insert_result(conn, row)
    conn.commit()


def _result_columns(conn: sqlite3.Connection) -> list[str]:
    info = conn.execute("PRAGMA table_info(experiment_results)").fetchall()
    return [str(row["name"]) for row in info]


def export_run_to_csv(conn: sqlite3.Connection, run_id: str, output_dir: Path | str) -> Path:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{run_id}.csv"
    cols = _result_columns(conn)
    rows = conn.execute(
        "SELECT * FROM experiment_results WHERE run_id = ? ORDER BY id ASC",
        (run_id,),
    ).fetchall()
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(cols)
        for row in rows:
            writer.writerow([row[col] for col in cols])
    return out_path


def export_run_to_json(conn: sqlite3.Connection, run_id: str, output_dir: Path | str) -> Path:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{run_id}.json"
    cols = _result_columns(conn)
    rows = conn.execute(
        "SELECT * FROM experiment_results WHERE run_id = ? ORDER BY id ASC",
        (run_id,),
    ).fetchall()
    serialisable = [{col: row[col] for col in cols} for row in rows]
    payload = {"run_id": run_id, "row_count": len(serialisable), "results": serialisable}
    with out_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2)
    return out_path


def export_run_summary_excel(conn: sqlite3.Connection, run_id: str, output_dir: Path | str) -> Path:
    """
    Export a multi-sheet workbook with run-level summaries.
    Sheets:
      1) run_metadata
      2) condition_summary
      3) prompt_type_summary
      4) metric_averages
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from exc

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{run_id}.xlsx"

    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "run_metadata"
    ws_cond = wb.create_sheet("condition_summary")
    ws_prompt = wb.create_sheet("prompt_type_summary")
    ws_metrics = wb.create_sheet("metric_averages")

    run_row = conn.execute(
        """
        SELECT run_id, mode, model, started_at, completed_at
        FROM experiment_runs
        WHERE run_id = ?
        LIMIT 1
        """,
        (run_id,),
    ).fetchone()
    totals = conn.execute(
        """
        SELECT
            COUNT(*) AS total_rows,
            ROUND(AVG(score), 4) AS avg_score,
            ROUND(AVG(latency_ms), 2) AS avg_latency_ms
        FROM experiment_results
        WHERE run_id = ?
        """,
        (run_id,),
    ).fetchone()
    ws_meta.append(["field", "value"])
    ws_meta.append(["run_id", run_id])
    ws_meta.append(["mode", str(run_row["mode"]) if run_row else "unknown"])
    ws_meta.append(["model", str(run_row["model"]) if run_row and run_row["model"] else "unknown"])
    ws_meta.append(["started_at", str(run_row["started_at"]) if run_row else ""])
    ws_meta.append(["completed_at", str(run_row["completed_at"]) if run_row else ""])
    ws_meta.append(["total_rows", int(totals["total_rows"] or 0)])
    ws_meta.append(["avg_score", float(totals["avg_score"] or 0.0)])
    ws_meta.append(["avg_latency_ms", float(totals["avg_latency_ms"] or 0.0)])

    ws_cond.append(
        [
            "condition",
            "expected_total",
            "blocked_total",
            "detected_total",
            "legitimate_total",
            "false_detect_total",
            "dsr",
            "ddr",
            "fpr",
        ]
    )
    cond_rows = conn.execute(
        """
        SELECT
            condition,
            SUM(CASE WHEN expected_drift=1 THEN 1 ELSE 0 END) AS expected_total,
            SUM(CASE WHEN expected_drift=1 AND blocked=1 THEN 1 ELSE 0 END) AS blocked_total,
            SUM(CASE WHEN expected_drift=1 AND drift_detected=1 THEN 1 ELSE 0 END) AS detected_total,
            SUM(CASE WHEN expected_drift=0 THEN 1 ELSE 0 END) AS legitimate_total,
            SUM(CASE WHEN expected_drift=0 AND drift_detected=1 THEN 1 ELSE 0 END) AS false_detect_total
        FROM experiment_results
        WHERE run_id = ?
        GROUP BY condition
        ORDER BY condition ASC
        """,
        (run_id,),
    ).fetchall()
    for row in cond_rows:
        expected = int(row["expected_total"] or 0)
        blocked = int(row["blocked_total"] or 0)
        detected = int(row["detected_total"] or 0)
        legit = int(row["legitimate_total"] or 0)
        false_detect = int(row["false_detect_total"] or 0)
        ws_cond.append(
            [
                str(row["condition"]),
                expected,
                blocked,
                detected,
                legit,
                false_detect,
                round(blocked / expected if expected else 0.0, 4),
                round(detected / expected if expected else 0.0, 4),
                round(false_detect / legit if legit else 0.0, 4),
            ]
        )

    ws_prompt.append(
        [
            "prompt_type",
            "count",
            "blocked",
            "detected",
            "avg_score",
            "avg_latency_ms",
        ]
    )
    prompt_rows = conn.execute(
        """
        SELECT
            prompt_type,
            COUNT(*) AS count_total,
            SUM(CASE WHEN blocked=1 THEN 1 ELSE 0 END) AS blocked_total,
            SUM(CASE WHEN drift_detected=1 THEN 1 ELSE 0 END) AS detected_total,
            ROUND(AVG(score), 4) AS avg_score,
            ROUND(AVG(latency_ms), 2) AS avg_latency_ms
        FROM experiment_results
        WHERE run_id = ?
        GROUP BY prompt_type
        ORDER BY prompt_type ASC
        """,
        (run_id,),
    ).fetchall()
    for row in prompt_rows:
        ws_prompt.append(
            [
                str(row["prompt_type"]),
                int(row["count_total"] or 0),
                int(row["blocked_total"] or 0),
                int(row["detected_total"] or 0),
                float(row["avg_score"] or 0.0),
                float(row["avg_latency_ms"] or 0.0),
            ]
        )

    ws_metrics.append(["metric", "avg_value"])
    metrics = conn.execute(
        """
        SELECT
            ROUND(AVG(COALESCE(base_ids, ids_score, 0.0)), 4) AS avg_base_ids,
            ROUND(AVG(COALESCE(mahalanobis, 0.0)), 4) AS avg_mahalanobis,
            ROUND(AVG(COALESCE(kl_divergence, 0.0)), 4) AS avg_kl_divergence,
            ROUND(AVG(COALESCE(js_divergence, 0.0)), 4) AS avg_js_divergence,
            ROUND(AVG(COALESCE(wasserstein, 0.0)), 4) AS avg_wasserstein,
            ROUND(AVG(COALESCE(hellinger, 0.0)), 4) AS avg_hellinger,
            ROUND(AVG(COALESCE(tool_frequency, 0.0)), 4) AS avg_tool_frequency
        FROM experiment_results
        WHERE run_id = ?
        """,
        (run_id,),
    ).fetchone()
    ws_metrics.append(["base_ids", float(metrics["avg_base_ids"] or 0.0)])
    ws_metrics.append(["mahalanobis", float(metrics["avg_mahalanobis"] or 0.0)])
    ws_metrics.append(["kl_divergence", float(metrics["avg_kl_divergence"] or 0.0)])
    ws_metrics.append(["js_divergence", float(metrics["avg_js_divergence"] or 0.0)])
    ws_metrics.append(["wasserstein", float(metrics["avg_wasserstein"] or 0.0)])
    ws_metrics.append(["hellinger", float(metrics["avg_hellinger"] or 0.0)])
    ws_metrics.append(["tool_frequency", float(metrics["avg_tool_frequency"] or 0.0)])

    wb.save(out_path)
    return out_path

